import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side
from datetime import datetime
import win32com.client as win32
import os

def decimal_to_time(decimal):
    total_seconds = decimal * 24 * 3600
    hours = int(total_seconds // 3600)
    minutes = int((total_seconds % 3600) // 60)
    time_str = f"{hours:02}:{minutes:02}"
    return datetime.strptime(time_str, "%H:%M").strftime("%I:%M %p")
   
def calculate_minutes_late(time_in, expected_time_in):
    delta = time_in - expected_time_in
    return int(delta.total_seconds() // 60)

def modify_columns(wb, sheet_name, time_in_col, time_out_col, work_hours_col, minutes_late_col, legends_col):
    ws = wb[sheet_name]

    headers = [cell.value for cell in ws[1]]
    time_in_index = headers.index(time_in_col) + 1
    time_out_index = headers.index(time_out_col) + 1
    work_hours_index = headers.index(work_hours_col) + 1
    minutes_late_index = headers.index(minutes_late_col) + 1
    legends_index = headers.index(legends_col) + 1


    for row in ws.iter_rows(min_row=2, max_col=max(time_in_index, time_out_index, work_hours_index, minutes_late_index, legends_index)):
        cell_time_in = row[time_in_index - 1]
        cell_time_out = row[time_out_index - 1]
        cell_work_hours = row[work_hours_index - 1]
        cell_minutes_late = row[minutes_late_index - 1]
        legends = row[legends_index -1]
        legendsBOOL = False
 
            
        
        if isinstance(cell_time_in.value, (float, int)):
            cell_time_in.value = decimal_to_time(cell_time_in.value)
        if isinstance(cell_time_out.value, (float, int)):
            cell_time_out.value = decimal_to_time(cell_time_out.value)
        
        if cell_work_hours.value:
            try:
               
                if cell_time_in.value and cell_time_out.value:
                
                    time_in_string = str(cell_time_in.value)
                    time_in = datetime.strptime(time_in_string,  "%I:%M %p")
                    formatted_time_in = time_in.strftime("%I:%M %p")
                    print(formatted_time_in)
                    
                    time_out_string = str(cell_time_out.value)
                    time_out = datetime.strptime(time_out_string, "%I:%M %p")
                    formatted_time_out = time_out.strftime("%I:%M %p")
                    
                    print("cell work hours value", cell_work_hours.value)
                    if legends.value == "RGOT" or legends.value == "RDOT" or legends.value == "LHOT" or legends.value == "SHOT":
                        cell_minutes_late.value = 0 
                        legendsBOOL = True
                        print("Legends Index:",legends_index)
                    elif cell_work_hours.value == 9 and legendsBOOL == False:
                        if time_out.strftime("%p") == "AM":
                            eight_pm = datetime.strptime("08:00 PM", "%I:%M %p")
                            formatted_eight_am = eight_pm.strftime("%I:%M %p")
                            if formatted_time_in < formatted_eight_am:
                                cell_minutes_late.value = 0
                            else:
                                cell_minutes_late.value = calculate_minutes_late(time_in, eight_pm)
                                
                        elif time_out.strftime("%p") == "PM":
                            eight_am = datetime.strptime("08:00 AM", "%I:%M %p")
                            formatted_eight_am = eight_am.strftime("%I:%M %p")
                            if formatted_time_in < formatted_eight_am:
                                cell_minutes_late.value = 0
                            else:
                                cell_minutes_late.value = calculate_minutes_late(time_in, eight_am)

                    elif cell_work_hours.value == 12 and legendsBOOL == False:
                        
                        if time_in.strftime("%p") == "AM":
                        ### If shift is 7:00 AM - 7:00 PM 
                            seven_am = datetime.strptime("07:00 AM", "%I:%M %p")
                            formatted_seven_am = seven_am.strftime("%I:%M %p")
                            if formatted_time_in < formatted_seven_am:
                                cell_minutes_late.value = 0
                            else:
                                cell_minutes_late.value = calculate_minutes_late(time_in, seven_am)

                        elif time_in.strftime("%p") == "PM":
                            seven_pm = datetime.strptime("07:00 PM", "%I:%M %p")
                            formatted_seven_pm = seven_pm.strftime("%I:%M %p")
                            if formatted_time_in < formatted_seven_pm:
                                cell_minutes_late.value = 0
                            else:
                                cell_minutes_late.value = calculate_minutes_late(time_in, seven_pm)
                    elif cell_work_hours.value > 8 and legendsBOOL == False:
                        
                        cell_minutes_late.value = 0
                        
            except ValueError as e:
                
                pass

def process_excel_files(input_file_path, output_file_path, sheet_name, time_in_col, time_out_col, work_hours_col, minutes_late_col, legends_col):
    excel = win32.Dispatch('Excel.Application')
    excel.Visible = False

    # Open the .xls file
    wb_xls = excel.Workbooks.Open(os.path.abspath(input_file_path))
    
    # Save it as .xlsx
    xlsx_temp_file = os.path.abspath('temp_converted.xlsx')
    wb_xls.SaveAs(xlsx_temp_file, FileFormat=51)  # 51 is the file format for .xlsx
    
    # Close the .xls workbook
    wb_xls.Close()
    
    # Open the .xlsx file with openpyxl
    wb_xlsx = openpyxl.load_workbook(xlsx_temp_file)
    
    # Modify columns
    modify_columns(wb_xlsx, sheet_name, time_in_col, time_out_col, work_hours_col, minutes_late_col, legends_col)
    
    # Save final output to the specified path
    wb_xlsx.save(output_file_path)
    
    # Clean up
    os.remove(xlsx_temp_file)
    excel.Quit()

    print(f"Processing complete. Saved as {output_file_path}")

# Usage example
input_file_path = 'SAMPLE - WITHOUT DATA.xls'
output_file_path = 'Out.xlsx'
sheet_name = 'Timekeep'
time_in_col = 'TIME IN'
time_out_col = 'TIME OUT'
work_hours_col = 'WORK HOURS'
minutes_late_col = 'MINUTES LATE'
legends_col = 'Legend'

process_excel_files(input_file_path, output_file_path, sheet_name, time_in_col, time_out_col, work_hours_col, minutes_late_col, legends_col)
