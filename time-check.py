# -*- coding: utf-8 -*-
import pythoncom
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side
from datetime import datetime
import win32com.client as win32
import os
import sys
import tkinter as tk
from tkinter import filedialog, messagebox, StringVar, ttk, Label
import customtkinter as ctk
from PIL import Image, ImageTk, ImageDraw
from PIL import Image, ImageTk, ImageDraw

def decimal_to_time(decimal):
    total_seconds = decimal * 24 * 3600
    hours = int(total_seconds // 3600)
    minutes = int((total_seconds % 3600) // 60)
    time_str = f"{hours:02}:{minutes:02}"
    return datetime.strptime(time_str, "%H:%M").strftime("%I:%M %p")

def calculate_minutes_late(actual_time_str, expected_time_str):
    try:
        delta = actual_time_str - expected_time_str
        return int(delta.total_seconds() // 60) # Ensure the difference is not negative
    except ValueError:
        # If the time string is empty or not in the expected format, return 0 minutes late
        return 0


def modify_columns(wb, sheet_name, time_in_col, time_out_col, work_hours_col, minutes_late_col, legends_col):
    ws = wb[sheet_name]

    headers = [cell.value for cell in ws[1]]
    time_in_index = headers.index(time_in_col) + 1
    time_out_index = headers.index(time_out_col) + 1
    work_hours_index = headers.index(work_hours_col) + 1
    minutes_late_index = headers.index(minutes_late_col) + 1
    legends_index = headers.index(legends_col) + 1


    for row in ws.iter_rows(min_row=2, max_col=max(time_in_index, time_out_index, work_hours_index, minutes_late_index, legends_index)):
        cell_time_in = row[time_in_index - 1]
        cell_time_out = row[time_out_index - 1]
        cell_work_hours = row[work_hours_index - 1]
        cell_minutes_late = row[minutes_late_index - 1]
        legends = row[legends_index -1]
        legendsBOOL = False
 
            
        
        if isinstance(cell_time_in.value, (float, int)):
            cell_time_in.value = decimal_to_time(cell_time_in.value)
        if isinstance(cell_time_out.value, (float, int)):
            cell_time_out.value = decimal_to_time(cell_time_out.value)
        
        if cell_work_hours.value:
            try:
               
                if cell_time_in.value and cell_time_out.value:
                
                    time_in_string = str(cell_time_in.value)
                    time_in = datetime.strptime(time_in_string,  "%I:%M %p")
                    formatted_time_in = time_in.strftime("%I:%M %p")
                    print(formatted_time_in)
                    
                    time_out_string = str(cell_time_out.value)
                    time_out = datetime.strptime(time_out_string, "%I:%M %p")
                    formatted_time_out = time_out.strftime("%I:%M %p")
                    
                    print("cell work hours value", cell_work_hours.value)
                    if legends.value == "RGOT" or legends.value == "RDOT" or legends.value == "LHOT" or legends.value == "SHOT":
                        cell_minutes_late.value = 0 
                        legendsBOOL = True
                        print("Legends Index:",legends_index)
                    elif cell_work_hours.value == 9 and legendsBOOL == False:
                        if time_out.strftime("%p") == "AM":
                            eight_pm = datetime.strptime("08:00 PM", "%I:%M %p")
                            formatted_eight_am = eight_pm.strftime("%I:%M %p")
                            if formatted_time_in < formatted_eight_am:
                                cell_minutes_late.value = 0
                            else:
                                cell_minutes_late.value = calculate_minutes_late(time_in, eight_pm)
                                
                        elif time_out.strftime("%p") == "PM":
                            eight_am = datetime.strptime("08:00 AM", "%I:%M %p")
                            formatted_eight_am = eight_am.strftime("%I:%M %p")
                            if formatted_time_in < formatted_eight_am:
                                cell_minutes_late.value = 0
                            else:
                                cell_minutes_late.value = calculate_minutes_late(time_in, eight_am)

                    elif cell_work_hours.value == 12 and legendsBOOL == False:
                        
                        if time_in.strftime("%p") == "AM":
                        ### If shift is 7:00 AM - 7:00 PM 
                            seven_am = datetime.strptime("07:00 AM", "%I:%M %p")
                            formatted_seven_am = seven_am.strftime("%I:%M %p")
                            if formatted_time_in < formatted_seven_am:
                                cell_minutes_late.value = 0
                            else:
                                cell_minutes_late.value = calculate_minutes_late(time_in, seven_am)

                        elif time_in.strftime("%p") == "PM":
                            seven_pm = datetime.strptime("07:00 PM", "%I:%M %p")
                            formatted_seven_pm = seven_pm.strftime("%I:%M %p")
                            if formatted_time_in < formatted_seven_pm:
                                cell_minutes_late.value = 0
                                
                            else:
                                cell_minutes_late.value = calculate_minutes_late(time_in, seven_pm)
                    elif cell_work_hours.value > 8 and legendsBOOL == False:
                        cell_minutes_late.value = 0
            
            except ValueError as e:
                
                pass
                
def process(input_file_path, output_file_path, sheet_name, time_in_col, time_out_col, work_hours_col, minutes_late_col, legends_col):
    excel = win32.Dispatch('Excel.Application')
    excel.Visible = False
    wb_xls = excel.Workbooks.Open(os.path.abspath(input_file_path))
    xlsx_temp_file = os.path.abspath('temp_converted.xlsx')
    wb_xls.SaveAs(xlsx_temp_file, FileFormat=51) 
    wb_xls.Close()
   
    wb_xlsx = openpyxl.load_workbook(xlsx_temp_file)
   
    modify_columns(wb_xlsx, sheet_name, time_in_col, time_out_col, work_hours_col, minutes_late_col, legends_col)
    wb_xlsx.save(output_file_path)
    os.remove(xlsx_temp_file)
    excel.Quit()

    print(f"Conversion complete. Saved as {output_file_path}")

def process_excel_files(input_file_path, output_file_path):
    sheet_name = 'Timekeep'
    time_in_col = 'TIME IN'
    time_out_col = 'TIME OUT'
    work_hours_col = 'WORK HOURS'
    minutes_late_col = 'MINUTES LATE'
    legends_col = 'Legend'
    process(input_file_path, output_file_path, sheet_name, time_in_col, time_out_col, work_hours_col, minutes_late_col, legends_col)


def select_input_file():
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xls;*.xlsx")])
    input_entry.delete(0, ctk.END)
    input_entry.insert(0, file_path)

def select_output_file(input_file_path):
    file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
    if file_path:
        try:
            process_excel_files(input_file_path, file_path)
            messagebox.showinfo("Success", f"File processed successfully and saved as {file_path}")
        except Exception as e:
            messagebox.showerror("Error", str(e))

def start_processing():
    input_file_path = input_entry.get()

    if not input_file_path:
        messagebox.showerror("Error", "Please select an input file")
        return

    select_output_file(input_file_path)

def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)

def time_diff_calculator():
    calculator_window = tk.Toplevel(app)  # Create a new window
    calculator_window.title("Time Diff Calculator")
    
    # Set dimensions for the calculator window
    window_width = 400
    window_height = 300
    calculator_window.geometry(f"{window_width}x{window_height}")

    # Get the position of the main window
    main_window_width = app.winfo_width()
    main_window_height = app.winfo_height()
    main_window_x = app.winfo_x()
    main_window_y = app.winfo_y()
    
    # Calculate the center position for the new window
    position_x = main_window_x + (main_window_width - window_width) // 2
    position_y = main_window_y + (main_window_height - window_height) // 2
    
    calculator_window.geometry(f"{window_width}x{window_height}+{position_x}+{position_y}")

    def calculate_time_difference():
        try:
            time1 = time1_entry.get()
            time1_hour, time1_minute = map(int, time1.split(":"))
            time1_period = time1_period_var.get()
            if time1_period == "PM" and time1_hour != 12:
                time1_hour += 12
            elif time1_period == "AM" and time1_hour == 12:
                time1_hour = 0

            time2 = time2_entry.get()
            time2_hour, time2_minute = map(int, time2.split(":"))
            time2_period = time2_period_var.get()
            if time2_period == "PM" and time2_hour != 12:
                time2_hour += 12
            elif time2_period == "AM" and time2_hour == 12:
                time2_hour = 0

            time1_total_minutes = time1_hour * 60 + time1_minute
            time2_total_minutes = time2_hour * 60 + time2_minute

            minute_difference = abs(time2_total_minutes - time1_total_minutes)

            result_label.configure(text=f"The minute difference is: {minute_difference} minutes")
        except ValueError:
            messagebox.showerror("Error", "Please enter valid time values")

    # Create a main frame to hold all the widgets
    main_frame = ctk.CTkFrame(calculator_window, fg_color="white")
    main_frame.pack(padx=20, pady=20, fill="both", expand=True)

    # Create a title label
    title_label = ctk.CTkLabel(main_frame, text="Time Calculator", font=ctk.CTkFont(size=24, weight="bold"), text_color="#000000")
    title_label.pack(pady=10)

    # Create a frame for time 1
    time1_frame = ctk.CTkFrame(main_frame, fg_color="white")
    time1_frame.pack(pady=10)

    time1_label = ctk.CTkLabel(time1_frame, text="Time 1:", font=ctk.CTkFont(size=16), text_color="#000000")
    time1_label.pack(side=tk.LEFT, padx=5)

    time1_entry = ctk.CTkEntry(time1_frame, width=60)
    time1_entry.pack(side=tk.LEFT, padx=5)

    time1_period_var = StringVar()
    time1_period_menu = ttk.Combobox(time1_frame, textvariable=time1_period_var, width=15)
    time1_period_menu['values'] = ("AM", "PM")
    time1_period_menu.current(0)
    time1_period_menu.pack(side=tk.LEFT, padx=5)

    # Create a frame for time 2
    time2_frame = ctk.CTkFrame(main_frame, fg_color="white")
    time2_frame.pack(pady=10)

    time2_label = ctk.CTkLabel(time2_frame, text="Time 2:", font=ctk.CTkFont(size=16), text_color="#000000")
    time2_label.pack(side=tk.LEFT, padx=5)

    time2_entry = ctk.CTkEntry(time2_frame, width=60)
    time2_entry.pack(side=tk.LEFT, padx=5)

    time2_period_var = StringVar()
    time2_period_menu = ttk.Combobox(time2_frame, textvariable=time2_period_var, width=15)
    time2_period_menu['values'] = ("AM", "PM")
    time2_period_menu.current(0)
    time2_period_menu.pack(side=tk.LEFT, padx=5)

    # Create a calculate button
    calculate_button = ctk.CTkButton(main_frame, text="Calculate", font=ctk.CTkFont(size=16, weight="bold"), command=calculate_time_difference)
    calculate_button.pack(pady=10)

    # Create a result label
    result_label = ctk.CTkLabel(main_frame, text="", font=ctk.CTkFont(size=16), text_color="#000000")
    result_label.pack(pady=10)

# Main application window


app = ctk.CTk()
app.geometry("1000x600")
app.title("Time Processor App")

image_path = resource_path("imgs/time.png")
img = Image.open(image_path)

background = Image.new("RGBA", img.size, (255, 255, 255, 0))  # Transparent background
background.paste(img, (0, 0), img)
background = background.convert("RGB")

img = background.resize((200, 200), Image.LANCZOS)
img = ImageTk.PhotoImage(img)
img_label = Label(app, image=img, bg='white', borderwidth=0, highlightthickness=0)
img_label.pack(pady=10)

frame = ctk.CTkFrame(app, width=900, height=500, corner_radius=10, fg_color="white")
frame.place(relx=0.5, rely=0.5, anchor=tk.CENTER)

label_title = ctk.CTkLabel(frame, text="HR Data Processor", font=ctk.CTkFont(size=24, weight="bold"), text_color="#000000")
label_title.pack(pady=20)

input_frame = ctk.CTkFrame(frame, width=800, height=100, fg_color="white")
input_frame.pack(pady=10)

input_label = ctk.CTkLabel(input_frame, text="Input Excel File:", font=ctk.CTkFont(size=16), text_color="#000000")
input_label.pack(side=tk.LEFT, padx=10)

input_entry = ctk.CTkEntry(input_frame, width=400)
input_entry.pack(side=tk.LEFT, padx=10)

input_button = ctk.CTkButton(input_frame, text="Browse", command=select_input_file, fg_color="#3498db")
input_button.pack(side=tk.LEFT, padx=10)

process_button = ctk.CTkButton(frame, text="Process", command=start_processing, fg_color="#2ecc71", font=ctk.CTkFont(size=16, weight="bold"))
process_button.pack(pady=20)

# Time Diff Calculator button
time_diff_button = ctk.CTkButton(frame, text="Time Diff Calculator", command=time_diff_calculator, fg_color="#e74c3c", font=ctk.CTkFont(size=16, weight="bold"))
time_diff_button.pack(pady=10)

app.mainloop()